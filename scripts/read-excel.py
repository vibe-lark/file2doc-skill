#!/usr/bin/env python3
# =============================================================================
# Excel/CSV 读取转 Markdown 表格脚本
# =============================================================================
# 用途: 读取 Excel (.xlsx/.xls) 或 CSV 文件，将每个 Sheet 转为 Markdown 表格
# 依赖: openpyxl (处理 xlsx)，缺少时自动安装
# 用法: python3 read-excel.py <Excel/CSV文件> [--max-rows N(默认无限制)]
# 示例: python3 read-excel.py data.xlsx
# 输出: Markdown 表格内容输出到 stdout
# 注意: 日志输出到 stderr，结果数据输出到 stdout
# =============================================================================

import sys
import os
import csv
import subprocess
import argparse


def log_info(msg):
    """日志输出到 stderr"""
    sys.stderr.write(f"[INFO] {msg}\n")


def log_error(msg):
    """错误输出到 stderr"""
    sys.stderr.write(f"[ERROR] {msg}\n")


def ensure_openpyxl():
    """自动检测并安装 openpyxl"""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        log_info("openpyxl 未安装，正在自动安装...")
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "--quiet", "openpyxl"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        log_info("openpyxl 安装完成")


def parse_args():
    """解析命令行参数"""
    parser = argparse.ArgumentParser(
        description="Excel/CSV 转 Markdown 表格"
    )
    parser.add_argument("file_path", help="Excel 或 CSV 文件路径")
    parser.add_argument(
        "--max-rows", type=int, default=0,
        help="每个 Sheet 最大读取行数，0 表示不限制 (默认: 0)"
    )
    return parser.parse_args()


def cell_to_str(value):
    """
    将单元格值转为安全的字符串
    - 处理 None 值
    - 转义 Markdown 管道符
    - 去除换行符
    """
    if value is None:
        return ""
    text = str(value).strip()
    # 转义管道符，避免破坏 Markdown 表格结构
    text = text.replace("|", "\\|")
    # 将换行替换为空格，保持表格单行
    text = text.replace("\n", " ").replace("\r", "")
    return text


def rows_to_markdown(rows, sheet_name=None):
    """
    将二维数据转为 Markdown 表格字符串
    - rows: 二维列表，第一行视为表头
    - sheet_name: 可选的 Sheet 名称，用于标题
    返回 Markdown 字符串
    """
    if not rows or len(rows) == 0:
        return ""

    lines = []

    # Sheet 标题
    if sheet_name:
        lines.append(f"## {sheet_name}")
        lines.append("")

    # 确定列数 (取所有行中最大列数)
    max_cols = max(len(row) for row in rows)
    if max_cols == 0:
        return ""

    # 规范化行：每行补齐到 max_cols 列
    normalized = []
    for row in rows:
        cells = [cell_to_str(c) for c in row]
        while len(cells) < max_cols:
            cells.append("")
        normalized.append(cells)

    # 表头行
    header = normalized[0]
    # 如果表头全空，生成默认列名
    if all(h == "" for h in header):
        header = [f"Col_{i+1}" for i in range(max_cols)]

    lines.append("| " + " | ".join(header) + " |")

    # 分隔行
    lines.append("| " + " | ".join(["---"] * max_cols) + " |")

    # 数据行
    for row in normalized[1:]:
        lines.append("| " + " | ".join(row) + " |")

    lines.append("")
    return "\n".join(lines)


def read_xlsx(file_path, max_rows):
    """
    读取 .xlsx 文件的所有 Sheet
    - file_path: 文件路径
    - max_rows: 每个 Sheet 最大行数, 0 表示不限
    返回 Markdown 字符串
    """
    import openpyxl

    log_info(f"正在读取 xlsx: {file_path}")
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        log_error(f"无法打开 Excel 文件: {e}")
        sys.exit(1)

    results = []
    sheet_names = wb.sheetnames
    log_info(f"共 {len(sheet_names)} 个 Sheet: {', '.join(sheet_names)}")

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        rows = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if max_rows > 0 and row_idx >= max_rows + 1:
                # +1 因为第一行是表头
                log_info(f"Sheet '{sheet_name}' 达到最大行数限制: {max_rows}")
                break
            rows.append(list(row))

        if rows:
            md = rows_to_markdown(rows, sheet_name)
            if md:
                results.append(md)
        else:
            log_info(f"Sheet '{sheet_name}' 为空，跳过")

    wb.close()
    return "\n".join(results)


def read_xls(file_path, max_rows):
    """
    读取 .xls 文件 (旧版 Excel 格式)
    尝试使用 openpyxl，如果失败则提示用户
    - file_path: 文件路径
    - max_rows: 每个 Sheet 最大行数, 0 表示不限
    返回 Markdown 字符串
    """
    log_info(f"正在读取 xls: {file_path}")
    log_info("注意: .xls 为旧版格式，建议转换为 .xlsx 后再处理")

    # 尝试用 openpyxl 打开 (部分 .xls 可以兼容)
    try:
        return read_xlsx(file_path, max_rows)
    except Exception:
        log_error(".xls 格式不受 openpyxl 支持，请先将文件转为 .xlsx 格式")
        log_error("可使用 LibreOffice 转换: soffice --headless --convert-to xlsx input.xls")
        sys.exit(1)


def read_csv(file_path, max_rows):
    """
    读取 CSV 文件
    - file_path: 文件路径
    - max_rows: 最大行数, 0 表示不限
    返回 Markdown 字符串
    """
    log_info(f"正在读取 CSV: {file_path}")

    rows = []
    # 自动检测编码和分隔符
    try:
        with open(file_path, "r", encoding="utf-8-sig") as f:
            # 使用 Sniffer 检测分隔符
            sample = f.read(8192)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = csv.excel

            reader = csv.reader(f, dialect)
            for row_idx, row in enumerate(reader):
                if max_rows > 0 and row_idx >= max_rows + 1:
                    log_info(f"达到最大行数限制: {max_rows}")
                    break
                rows.append(row)
    except UnicodeDecodeError:
        # 回退到 gbk 编码 (中文 Windows 常见)
        log_info("UTF-8 解码失败，尝试 GBK 编码...")
        with open(file_path, "r", encoding="gbk") as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader):
                if max_rows > 0 and row_idx >= max_rows + 1:
                    break
                rows.append(row)
    except Exception as e:
        log_error(f"无法读取 CSV 文件: {e}")
        sys.exit(1)

    if not rows:
        log_error("CSV 文件为空")
        sys.exit(1)

    log_info(f"读取 {len(rows)} 行数据")
    sheet_name = os.path.splitext(os.path.basename(file_path))[0]
    return rows_to_markdown(rows, sheet_name)


def main():
    args = parse_args()

    file_path = args.file_path
    max_rows = args.max_rows

    # 校验输入文件
    if not os.path.isfile(file_path):
        log_error(f"文件不存在: {file_path}")
        sys.exit(1)

    # 判断文件类型
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".xlsx":
        ensure_openpyxl()
        result = read_xlsx(file_path, max_rows)
    elif ext == ".xls":
        ensure_openpyxl()
        result = read_xls(file_path, max_rows)
    elif ext == ".csv":
        result = read_csv(file_path, max_rows)
    else:
        log_error(f"不支持的文件类型: {ext} (支持: .xlsx, .xls, .csv)")
        sys.exit(1)

    if not result.strip():
        log_error("未读取到有效数据")
        sys.exit(1)

    # 输出 Markdown 到 stdout
    print(result)


if __name__ == "__main__":
    main()
